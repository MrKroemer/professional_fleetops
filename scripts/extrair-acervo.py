#!/usr/bin/env python3
"""
Extrai o acervo das planilhas legadas para um arquivo JSON de seed.

Este script roda **fora** da aplicação, em tempo de desenvolvimento: ele lê os
arquivos de `arquivos/` e escreve `backend/src/main/resources/db/seed/acervo.json`,
que o seed do perfil `dev` carrega. O importador de verdade — com relatório de
linhas aceitas e rejeitadas (RN-24) — é entrega da Fase 5; aqui o objetivo é
apenas dar ao ambiente de desenvolvimento um volume de dados realista, cruzado
entre obras, locadoras, condutores, veículos, fornecedores e tabelas de preço.

Uso:  python3 scripts/extrair-acervo.py
"""

from __future__ import annotations

import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

RAIZ = Path(__file__).resolve().parent.parent
ARQUIVOS = RAIZ / "arquivos"
SAIDA = RAIZ / "backend/src/main/resources/db/seed/acervo.json"

import openpyxl  # noqa: E402  (import após as constantes, para falhar cedo com mensagem clara)


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

PADRAO_PLACA = re.compile(r"^[A-Z]{3}[0-9][A-Z0-9][0-9]{2}$")

UFS_VALIDAS = {
    "AC","AL","AM","AP","BA","CE","DF","ES","GO","MA","MG","MS","MT",
    "PA","PB","PE","PI","PR","RJ","RN","RO","RR","RS","SC","SE","SP","TO",
}

# Cidades que aparecem sem UF nas planilhas; o estado vem do contexto da obra.
UF_POR_CIDADE = {
    "MONTES CLAROS": "MG", "PICOS": "PI", "SALVADOR": "BA", "FORTALEZA": "CE",
    "BARREIRAS": "BA", "JUAZEIRO": "BA", "RECIFE": "PE", "NATAL": "RN",
    "FEIRA DE SANTANA": "BA", "VITORIA DA CONQUISTA": "BA", "GOIANIA": "GO",
    "RIO DE JANEIRO": "RJ", "BRASILIA": "DF", "GUARATINGUETA": "SP",
    "JUAZEIRO DO NORTE": "CE", "PONTA NEGRA": "RN", "IRECÊ": "BA", "IRECE": "BA",
    "UIBAÍ": "BA", "UIBAI": "BA", "SEABRA": "BA", "PIRAPORA": "MG",
    "BROTAS DE MACAÚBAS": "BA", "FERNANDO DE NORONHA": "PE", "RUSSAS": "CE",
    "CAJAZEIRAS": "PB", "PENTECOSTE": "CE", "IBIMIRIM": "PE", "ARINOS": "MG",
    "HORIZONTE": "CE", "NOVA OLINDA": "CE", "ANGRA DOS REIS": "RJ",
}


def texto(valor) -> str:
    return "" if valor is None else str(valor).strip()


def sem_acento(valor: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", valor) if unicodedata.category(c) != "Mn"
    )


def normalizar_placa(bruta: str) -> str | None:
    limpa = re.sub(r"[\s.\-]", "", texto(bruta)).upper()
    return limpa if PADRAO_PLACA.match(limpa) else None


def cidade_e_uf(local: str) -> tuple[str, str]:
    """Separa "MONTES CLAROS - MG" em cidade e UF, inferindo quando a UF falta."""
    bruto = texto(local).replace("AEROPORTO", "").strip(" -")
    if not bruto:
        return "", ""
    partes = re.split(r"\s*[-/]\s*", bruto)
    uf = ""
    if len(partes) > 1 and partes[-1].upper() in UFS_VALIDAS:
        uf = partes[-1].upper()
        cidade = " - ".join(partes[:-1]).strip()
    else:
        cidade = bruto
    chave = sem_acento(cidade).upper()
    if not uf:
        for nome, sigla in UF_POR_CIDADE.items():
            if sem_acento(nome).upper() in chave:
                uf = sigla
                break
    return cidade.title(), uf or "PE"


def cpf_sintetico(semente: int) -> str:
    """
    Gera um CPF válido e determinístico a partir de um índice.

    As planilhas não trazem CPF dos condutores. Como o campo é obrigatório e único
    no domínio, o seed usa números sinteticamente válidos — nunca CPFs reais.
    """
    base = f"{(semente * 7919 + 100000000) % 1000000000:09d}"
    if len(set(base)) == 1:
        base = base[:-1] + str((int(base[-1]) + 1) % 10)

    def digito(numero: str, peso_inicial: int) -> int:
        soma = sum(int(d) * (peso_inicial - i) for i, d in enumerate(numero))
        resto = soma % 11
        return 0 if resto < 2 else 11 - resto

    primeiro = digito(base, 10)
    segundo = digito(base + str(primeiro), 11)
    return f"{base}{primeiro}{segundo}"


def combustivel_de(modelo: str, categoria: str) -> str:
    """
    Infere o combustível a partir do modelo.

    As picapes de trabalho da frota (S10, L200, Hilux, Ranger) são a diesel — é o que
    torna obrigatório o teste de fumaça preta na retirada (RN-09). O restante é flex.
    """
    modelo = sem_acento(modelo).upper()
    if any(p in modelo for p in ("S10", "L200", "HILUX", "RANGER", "AMAROK", "FRONTIER")):
        return "DIESEL"
    return "FLEX"


def categoria_de(bruta: str, modelo: str) -> str:
    """
    Traduz a categoria da planilha para o domínio.

    A planilha usa a nomenclatura de tração da locadora: "4X2" designa os SUVs
    compactos (T-Cross, Tracker, Creta, Renegade) e "4X4" as picapes. O domínio
    trabalha com a categoria tarifária, que é o que determina o preço do KM
    excedente (RN-06).
    """
    valor = sem_acento(texto(bruta)).upper().replace(" 1.6", "").strip()
    modelo_limpo = sem_acento(texto(modelo)).upper()
    if valor == "4X4":
        return "QUATRO_X_QUATRO"
    if valor == "4X2":
        return "UTILITARIO" if "TORO" in modelo_limpo or "STRADA" in modelo_limpo else "SUV"
    if valor.startswith("PASSEIO"):
        return "PASSEIO"
    return "PASSEIO"


def moeda(valor) -> float | None:
    """Interpreta valores que aparecem ora como número, ora como "R$ 2.805,30"."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return round(float(valor), 2)
    limpo = re.sub(r"[^\d,.-]", "", str(valor)).strip()
    if not limpo:
        return None
    if "," in limpo:
        limpo = limpo.replace(".", "").replace(",", ".")
    try:
        return round(float(limpo), 2)
    except ValueError:
        return None


def abrir(nome: str):
    caminho = ARQUIVOS / nome
    if not caminho.exists():
        raise SystemExit(f"Planilha não encontrada: {caminho}")
    return openpyxl.load_workbook(caminho, read_only=True, data_only=True)


def linhas_com_cabecalho(ws, linha_do_cabecalho: int):
    linhas = list(ws.iter_rows(min_row=linha_do_cabecalho, values_only=True))
    cabecalho = [texto(c) for c in linhas[0]]
    return [dict(zip(cabecalho, linha)) for linha in linhas[1:]]


# ---------------------------------------------------------------------------
# Extração
# ---------------------------------------------------------------------------

DIAS = {
    "SEG": ["SEGUNDA", "SEG"], "TER": ["TERCA", "TER"], "QUA": ["QUARTA", "QUA"],
    "QUI": ["QUINTA", "QUI"], "SEX": ["SEXTA", "SEX"], "SAB": ["SABADO", "SAB"],
    "DOM": ["DOMINGO", "DOM"],
}
ORDEM_DIAS = list(DIAS)


def dias_autorizados(descricao: str) -> list[str]:
    """
    Interpreta "terça, quinta e sábado" ou "terça à sábado" (RN-04).

    Um intervalo com "à" vira a sequência completa entre os extremos, que é como o
    gestor lê a anotação da planilha.
    """
    bruto = sem_acento(texto(descricao)).upper()
    if not bruto or "CONFIRMAR" in bruto or "DESATIVADO" in bruto:
        return []
    encontrados = [sigla for sigla, nomes in DIAS.items() if any(n in bruto for n in nomes)]
    if len(encontrados) == 2 and (" A " in bruto or " AH " in bruto or "À" in texto(descricao).upper()):
        inicio, fim = ORDEM_DIAS.index(encontrados[0]), ORDEM_DIAS.index(encontrados[1])
        if inicio <= fim:
            return ORDEM_DIAS[inicio : fim + 1]
    return [d for d in ORDEM_DIAS if d in encontrados]


def extrair_contratos() -> list[dict]:
    wb = abrir("FOR.FRO.06_Controle geral de veículos.xlsx")
    registros = linhas_com_cabecalho(wb["Locação Mensal"], 3)
    wb.close()
    return [r for r in registros if texto(r.get("PLACA"))]


def extrair_obras(contratos: list[dict]) -> list[dict]:
    """Uma obra por código, com o nome mais frequente e a localização mais comum."""
    nomes: dict[str, Counter] = defaultdict(Counter)
    locais: dict[str, Counter] = defaultdict(Counter)
    ativas: set[str] = set()

    for linha in contratos:
        codigo_bruto = texto(linha.get("CÓDIGO"))
        if not codigo_bruto or codigo_bruto in {"SEDE", "-"}:
            codigo_bruto = codigo_bruto or ""
        if not codigo_bruto:
            continue
        codigo = formatar_codigo_de_obra(codigo_bruto)
        nome = texto(linha.get("OBRA"))
        if nome:
            nomes[codigo][nome] += 1
        local = texto(linha.get("LOCAL DE RETIRADA")) or texto(linha.get("LOCAL"))
        if local:
            locais[codigo][local] += 1
        if texto(linha.get("STATUS")).upper() == "ATIVO":
            ativas.add(codigo)

    obras = []
    for codigo in sorted(nomes):
        nome = nomes[codigo].most_common(1)[0][0]
        cidade, uf = cidade_e_uf(locais[codigo].most_common(1)[0][0] if locais[codigo] else "")
        obras.append({
            "codigo": codigo,
            "nome": nome.title() if nome.isupper() else nome,
            "cliente": cliente_de(nome),
            "cidade": cidade or "Recife",
            "uf": uf,
            "status": "ATIVA" if codigo in ativas else "ENCERRADA",
        })
    return obras


def formatar_codigo_de_obra(bruto: str) -> str:
    """Converte 24019 no formato 24.019 usado no dia a dia."""
    digitos = re.sub(r"\D", "", bruto)
    if len(digitos) == 5:
        return f"{digitos[:2]}.{digitos[2:]}"
    return bruto.upper()


def cliente_de(nome_da_obra: str) -> str | None:
    """O nome da obra costuma trazer o cliente como prefixo."""
    conhecidos = [
        "STATKRAFT", "CGN", "AUREN", "SKER", "ELERA", "KROMA", "MINGYANG",
        "TOTAL EREN", "FURNAS", "DNIT", "CASA DOS VENTOS", "ALYA",
    ]
    alvo = sem_acento(nome_da_obra).upper()
    for cliente in conhecidos:
        if cliente in alvo:
            return cliente.title()
    return None


def extrair_locadoras() -> list[dict]:
    wb = abrir("Controle Geral - Fornecedores Frotas.xlsx")
    locadoras: list[dict] = []

    for linha in linhas_com_cabecalho(wb["LOCADORAS"], 2):
        nome = texto(linha.get("Empresa"))
        if not nome:
            continue
        locadoras.append({
            "nome": nome,
            "tipo": "NACIONAL",
            "consultor": texto(linha.get("Consultor (a)")) or None,
            "telefone": texto(linha.get("Contato")) or None,
            "email": primeiro_email(texto(linha.get("E-mail"))),
            "portalUrl": texto(linha.get("Portal"))[:400] or None,
            # As planilhas guardam login e senha em texto claro; aqui eles entram
            # cifrados (RN-20). Os valores abaixo são de demonstração.
            "portalLogin": "proyfebrasil" if texto(linha.get("Portal")) else None,
            "portalSenha": "portal@2026" if texto(linha.get("Portal")) else None,
            "canais": {
                "reservas": texto(linha.get("Reservas"))[:200] or None,
                "manutencao": texto(linha.get("Manutenção"))[:200] or None,
                "guinchoSinistro": texto(linha.get("Guincho/Sinistro"))[:200] or None,
                "assistencia24h": texto(linha.get("Assistência 24h"))[:200] or None,
                "financeiro": texto(linha.get("Financeiro"))[:200] or None,
                "suporte": texto(linha.get("Suporte ao Cliente"))[:200] or None,
                "telemetria": texto(linha.get("Telemetria"))[:200] or None,
            },
            "observacoes": texto(linha.get("Observações"))[:2000] or None,
            "ativa": True,
        })

    for linha in linhas_com_cabecalho(wb["LOCADORAS AVULSAS"], 2):
        nome = texto(linha.get("Empresa"))
        if not nome:
            continue
        locadoras.append({
            "nome": nome,
            "tipo": "AVULSA",
            "consultor": texto(linha.get("Consultor (a)")) or None,
            "telefone": texto(linha.get("Contato")) or None,
            "email": None,
            "portalUrl": None,
            "portalLogin": None,
            "portalSenha": None,
            "canais": {"assistencia24h": texto(linha.get("Assistencia"))[:200] or None},
            "observacoes": " · ".join(
                filter(None, [texto(linha.get("Local")), texto(linha.get("Forma de pagamento"))])
            )[:2000] or None,
            "ativa": True,
        })

    wb.close()

    # Locadoras que só aparecem no controle de veículos.
    conhecidas = {sem_acento(loc["nome"]).upper() for loc in locadoras}
    for extra in ("SpeedWay", "Veículo do profissional", "Agregado"):
        if sem_acento(extra).upper() not in conhecidas:
            locadoras.append({
                "nome": extra,
                "tipo": "AVULSA",
                "consultor": None, "telefone": None, "email": None, "portalUrl": None,
                "portalLogin": None, "portalSenha": None, "canais": {},
                "observacoes": "Registrada a partir do controle geral de veículos.",
                "ativa": True,
            })
    return locadoras


def primeiro_email(bruto: str) -> str | None:
    achado = re.search(r"[\w.+-]+@[\w-]+\.[\w.]+", bruto)
    return achado.group(0) if achado else None


def extrair_condutores(contratos: list[dict], obras: list[dict]) -> list[dict]:
    """Um condutor por nome distinto, alocado na obra em que aparece mais vezes."""
    por_obra: dict[str, Counter] = defaultdict(Counter)
    ativos: set[str] = set()
    for linha in contratos:
        nome = texto(linha.get("USUARIO"))
        if not nome or nome.upper() in {"-", "N/A"}:
            continue
        codigo = formatar_codigo_de_obra(texto(linha.get("CÓDIGO")))
        if codigo:
            por_obra[nome][codigo] += 1
        if texto(linha.get("STATUS")).upper() == "ATIVO":
            ativos.add(nome)

    codigos_validos = {o["codigo"] for o in obras}
    condutores = []
    for indice, nome in enumerate(sorted(por_obra), start=1):
        codigo = por_obra[nome].most_common(1)[0][0]
        condutores.append({
            "nome": nome.title() if nome.isupper() else nome,
            "cpf": cpf_sintetico(indice),
            "cargo": None,
            "obraCodigo": codigo if codigo in codigos_validos else None,
            "status": "ATIVO" if nome in ativos else "INATIVO",
            # A validade da CNH não existe nas planilhas. É distribuída de forma
            # determinística para que a tela de alertas (RN-16) tenha o que mostrar.
            "cnhValidadeDeslocamentoEmDias": (indice * 37) % 900 - 120,
            "cnhCategoria": "AB" if indice % 4 == 0 else "B",
        })
    return condutores


COLUNAS_DE_SUBSTITUICAO = [
    ("DATA 1ª SUBS", "MODELO"),
    ("DATA 2ª SUBS", "MODELO2"),
    ("DATA 3 SUBS", "MODELO3"),
    ("DATA 4ª SUBS", "MODELO4"),
    ("DATA 5ª SUBS2", "MODELO6"),
    ("DATA 6ª SUBS22", "MODELO7"),
]

PADRAO_PLACA_SOLTA = re.compile(r"[A-Z]{3}[0-9][A-Z0-9][0-9]{2}")


def substituicoes_da_linha(linha: dict) -> list[dict]:
    """
    Lê as substituições de veículo de uma linha do controle geral.

    A planilha registra até seis trocas em colunas repetidas — `MODELO`, `MODELO2`,
    até `MODELO7` —, cada uma com a data ao lado. O conteúdo da célula mistura modelo
    e placa em ordem livre ("TCROSS ENI9F24" e "RTX9C83 S10" aparecem os dois), então
    a placa é localizada por padrão e o que sobra vira o modelo.

    As entradas saem ordenadas por data, porque a planilha nem sempre respeita a
    sequência das colunas: há linhas em que a "2ª substituição" é anterior à primeira.
    """
    encontradas = []
    for coluna_data, coluna_modelo in COLUNAS_DE_SUBSTITUICAO:
        bruto = texto(linha.get(coluna_modelo))
        data = data_iso(linha.get(coluna_data))
        if not bruto or not data:
            continue
        limpo = re.sub(r"[\s.\-]", "", bruto).upper()
        achado = PADRAO_PLACA_SOLTA.search(limpo)
        if not achado:
            continue
        placa = achado.group(0)
        # O modelo é o texto sem a placa; sem isso o cadastro do veículo substituto
        # ficaria com o nome "TCROSS ENI9F24".
        modelo = re.sub(PADRAO_PLACA_SOLTA, "", limpo).strip() or "Não informado"
        encontradas.append({"data": data, "placa": placa, "modelo": modelo.title()})

    encontradas.sort(key=lambda item: item["data"])
    return encontradas


def extrair_veiculos(contratos: list[dict], locadoras: list[dict]) -> tuple[list[dict], list[str]]:
    nomes_de_locadora = {sem_acento(loc["nome"]).upper(): loc["nome"] for loc in locadoras}
    vistos: set[str] = set()
    veiculos: list[dict] = []
    rejeitadas: list[str] = []

    for linha in contratos:
        placa = normalizar_placa(texto(linha.get("PLACA")))
        if not placa:
            bruta = texto(linha.get("PLACA"))
            if bruta:
                rejeitadas.append(bruta)
            continue
        if placa in vistos:
            continue
        vistos.add(placa)

        modelo = texto(linha.get("CARRO")) or "Não informado"
        bruta_locadora = sem_acento(texto(linha.get("LOCADORA"))).upper()
        locadora = nomes_de_locadora.get(bruta_locadora)
        if not locadora:
            locadora = next(
                (nome for chave, nome in nomes_de_locadora.items() if chave and chave in bruta_locadora),
                None,
            )
        if not locadora:
            continue

        ativo = texto(linha.get("STATUS")).upper() == "ATIVO"
        # Um veículo que entrou por substituição foi devolvido quando o contrato
        # seguiu com outro; só o último período fica em uso.
        substituicoes = substituicoes_da_linha(linha)
        veiculos.append({
            "placa": placa,
            "modelo": modelo.title() if modelo.isupper() else modelo,
            "fabricante": None,
            "categoria": categoria_de(texto(linha.get("CATEGORIA")), modelo),
            "combustivel": combustivel_de(modelo, texto(linha.get("CATEGORIA"))),
            "locadoraNome": locadora,
            "grupoTarifario": None,
            "codigoInterno": formatar_codigo_de_obra(texto(linha.get("CÓDIGO"))) or None,
            "possuiRastreador": texto(linha.get("CARRO COM RASTREADOR")).upper().startswith("SIM"),
            "fornecedorRastreador": "Recife GPS"
                if texto(linha.get("CARRO COM RASTREADOR")).upper().startswith("SIM") else None,
            "possuiAdesivo": ativo,
            # O veículo da coluna PLACA só continua em uso se nunca foi substituído.
            "status": "EM_USO" if ativo and not substituicoes else "DEVOLVIDO",
        })

        for indice, troca in enumerate(substituicoes):
            if troca["placa"] in vistos:
                continue
            vistos.add(troca["placa"])
            ultima = indice == len(substituicoes) - 1
            veiculos.append({
                "placa": troca["placa"],
                "modelo": troca["modelo"],
                "fabricante": None,
                "categoria": categoria_de(texto(linha.get("CATEGORIA")), troca["modelo"]),
                "combustivel": combustivel_de(troca["modelo"], texto(linha.get("CATEGORIA"))),
                "locadoraNome": locadora,
                "grupoTarifario": None,
                "codigoInterno": formatar_codigo_de_obra(texto(linha.get("CÓDIGO"))) or None,
                "possuiRastreador": texto(linha.get("CARRO COM RASTREADOR")).upper().startswith("SIM"),
                "fornecedorRastreador": "Recife GPS"
                    if texto(linha.get("CARRO COM RASTREADOR")).upper().startswith("SIM") else None,
                "possuiAdesivo": ativo and ultima,
                "status": "EM_USO" if ativo and ultima else "DEVOLVIDO",
            })
    return veiculos, rejeitadas


def obras_por_codigo_na_celula(bruto: str, codigos: set[str]) -> list[str]:
    """As planilhas de fornecedores trazem "24.019 SKER Ventos..." na coluna Obra."""
    achados = re.findall(r"\d{2}\.?\d{3}", texto(bruto))
    resultado = []
    for achado in achados:
        codigo = formatar_codigo_de_obra(achado)
        if codigo in codigos:
            resultado.append(codigo)
    return resultado


def extrair_fornecedores(codigos_de_obra: set[str]) -> list[dict]:
    wb = abrir("Controle Geral - Fornecedores Frotas.xlsx")
    fornecedores: list[dict] = []
    vistos: set[tuple[str, str]] = set()

    def registrar(tipo: str, nome: str, **campos):
        chave = (tipo, sem_acento(nome).upper())
        if not nome or chave in vistos:
            return
        vistos.add(chave)
        fornecedores.append({"tipo": tipo, "nome": nome, **campos})

    for linha in linhas_com_cabecalho(wb["ABASTECIMENTO"], 2):
        cidade, uf = cidade_e_uf(texto(linha.get("Cidade")))
        registrar(
            "POSTO", texto(linha.get("Posto")),
            cidade=cidade or None, uf=uf or None,
            endereco=texto(linha.get("Endereço"))[:300] or None,
            telefone=texto(linha.get("Contato"))[:120] or None,
            email=primeiro_email(texto(linha.get("E-mail"))),
            responsavel=None,
            funcionamento=texto(linha.get("Funcionamento"))[:200] or None,
            formaFaturamento=texto(linha.get("Faturamento"))[:200] or None,
            formaPagamento=None,
            ativo=texto(linha.get("Status")).upper() != "INATIVO",
            observacoes=texto(linha.get("Observações"))[:2000] or None,
            obrasCodigos=obras_por_codigo_na_celula(linha.get("Obra"), codigos_de_obra),
            posto={
                "diasAutorizados": dias_autorizados(texto(linha.get("Dias Autorizados "))),
                "acessoFaturas": texto(linha.get("Acesso Faturas"))[:120] or None,
            },
        )

    for linha in linhas_com_cabecalho(wb["LAVA-JATO"], 2):
        cidade, uf = cidade_e_uf(texto(linha.get("Cidade")))
        ativo = "DESATIVAD" not in sem_acento(texto(linha.get("Responsável"))).upper()
        registrar(
            "LAVA_JATO", texto(linha.get("Lava Jato")),
            cidade=cidade or None, uf=uf or None,
            endereco=texto(linha.get("Endereço"))[:300] or None,
            telefone=texto(linha.get("Contato"))[:120] or None,
            email=primeiro_email(texto(linha.get("E-mail"))),
            responsavel=texto(linha.get("Responsável"))[:160] or None,
            funcionamento=texto(linha.get("Funcionamento"))[:200] or None,
            formaFaturamento=texto(linha.get("Faturamento"))[:200] or None,
            formaPagamento=texto(linha.get("Pagamento"))[:200] or None,
            ativo=ativo,
            observacoes=texto(linha.get("Observações"))[:2000] or None,
            obrasCodigos=obras_por_codigo_na_celula(linha.get("Obra"), codigos_de_obra),
            lavaJato={
                "servicosPorSemana": 1,
                "precoPasseio": moeda(linha.get("Passeio")),
                "precoSuv": moeda(linha.get("SUV")),
                "precoQuatroXQuatro": moeda(linha.get("4X4")),
            },
        )

    for aba, tipo, coluna in (("PNEU", "BORRACHARIA", "Oficina"), ("PARA-BRISAS", "PARA_BRISAS", "Oficina")):
        for linha in linhas_com_cabecalho(wb[aba], 2):
            cidade, uf = cidade_e_uf(texto(linha.get("Cidade")))
            registrar(
                tipo, texto(linha.get(coluna)),
                cidade=cidade or None, uf=uf or None,
                endereco=texto(linha.get("Endereço"))[:300] or None,
                telefone=texto(linha.get("Contato"))[:120] or None,
                email=primeiro_email(texto(linha.get("E-mail"))),
                responsavel=None,
                funcionamento=texto(linha.get("Funcionamento"))[:200] or None,
                formaFaturamento=None,
                formaPagamento=texto(linha.get("Pagamento"))[:200] or None,
                ativo=True,
                observacoes=texto(linha.get("Serviços"))[:2000] or None,
                obrasCodigos=obras_por_codigo_na_celula(linha.get("Obra"), codigos_de_obra),
            )

    for linha in linhas_com_cabecalho(wb["RASTREADOR"], 2):
        cidade, uf = cidade_e_uf(texto(linha.get("Cidade")))
        registrar(
            "RASTREADOR", texto(linha.get("Rastreio")),
            cidade=cidade or None, uf=uf or None,
            endereco=texto(linha.get("Endereço"))[:300] or None,
            telefone=texto(linha.get("Contato"))[:120] or None,
            email=primeiro_email(texto(linha.get("E-mail"))),
            responsavel=texto(linha.get("Responsável"))[:160] or None,
            funcionamento=None, formaFaturamento=None, formaPagamento=None,
            ativo=True,
            observacoes=texto(linha.get("Observações"))[:2000] or None,
            obrasCodigos=obras_por_codigo_na_celula(linha.get("Obra"), codigos_de_obra),
            rastreador={
                "mensalidade": moeda(linha.get("Mensalidade")),
                "custoInstalacao": moeda(linha.get("Instalação")),
                "custoDesinstalacao": moeda(linha.get("Desinstalação")),
                "equipadora": texto(linha.get("Equipadora"))[:180] or None,
                "portalUrl": texto(linha.get("Portal"))[:400] or None,
                "portalLogin": "proyfebrasil",
                "portalSenha": "telemetria@2026",
            },
        )

    for linha in linhas_com_cabecalho(wb["GRÁFICAS GERAL"], 2):
        cidade, uf = cidade_e_uf(texto(linha.get("Local")))
        registrar(
            "GRAFICA", texto(linha.get("Gráfica")),
            cidade=cidade or None, uf=uf or None,
            endereco=None,
            telefone=texto(linha.get("Contato"))[:120] or None,
            email=None, responsavel=None, funcionamento=None,
            formaFaturamento=None, formaPagamento=None,
            ativo=True,
            observacoes=texto(linha.get("Obs"))[:2000] or None,
            obrasCodigos=[],
            grafica={
                "tamanhoAdesivo": texto(linha.get("Adesivo"))[:40] or None,
                "precoAdesivo": moeda(linha.get("Valor")),
                "tamanhoIma": texto(linha.get("Imã"))[:40] or None,
                "precoIma": moeda(linha.get("Tamanho")),
            },
        )

    wb.close()
    return fornecedores


def data_iso(valor) -> str | None:
    """Converte a data da planilha para ISO, aceitando datetime ou texto."""
    if valor is None:
        return None
    if hasattr(valor, "date"):
        return valor.date().isoformat()
    bruto = texto(valor)
    achado = re.match(r"(\d{2})/(\d{2})/(\d{4})", bruto)
    if achado:
        return f"{achado.group(3)}-{achado.group(2)}-{achado.group(1)}"
    return None


def extrair_contratos_de_locacao(
    contratos: list[dict], codigos_de_obra: set[str], placas: set[str], condutores: list[dict]
) -> list[dict]:
    """
    Monta o contrato de locação de cada linha do controle geral.

    Cada linha da planilha é exatamente um contrato: uma obra, um condutor, um veículo,
    uma data de retirada e um pacote de KM. É esse vínculo que responde "quem dirige a
    placa X e em que obra" — pergunta que o cadastro de veículo, sozinho, não responde.
    """
    nomes_de_condutor = {sem_acento(c["nome"]).upper(): c["nome"] for c in condutores}
    resultado: list[dict] = []
    placas_usadas: set[str] = set()

    for linha in contratos:
        placa = normalizar_placa(texto(linha.get("PLACA")))
        codigo = formatar_codigo_de_obra(texto(linha.get("CÓDIGO")))
        if not placa or placa not in placas or codigo not in codigos_de_obra:
            continue
        # Uma placa pode reaparecer em linhas antigas; o contrato é o da primeira
        # ocorrência, que a planilha ordena da mais recente para a mais antiga.
        if placa in placas_usadas:
            continue
        placas_usadas.add(placa)

        bruto_condutor = sem_acento(texto(linha.get("USUARIO"))).upper()
        condutor = nomes_de_condutor.get(bruto_condutor)

        pacote_bruto = texto(linha.get("KM CONTRATADO")).upper()
        pacote = int(pacote_bruto) if pacote_bruto.isdigit() else None
        ativo = texto(linha.get("STATUS")).upper() == "ATIVO"

        resultado.append({
            "obraCodigo": codigo,
            "placa": placa,
            "condutorNome": condutor,
            "localRetirada": texto(linha.get("LOCAL DE RETIRADA")) or texto(linha.get("LOCAL")) or None,
            "dataRetirada": data_iso(linha.get("RETIRADA")),
            "dataEncerramento": data_iso(linha.get("DATA ENCERRAMENTO")),
            "pacoteKmContratado": pacote,
            # "ILIMITADO" e "LIVRE" aparecem em veículos do próprio profissional, que
            # não têm franquia contratada — o campo fica nulo e a observação registra.
            "observacoes": (
                f"Pacote registrado como \"{pacote_bruto}\" na planilha."
                if pacote is None and pacote_bruto else None
            ),
            "status": "ATIVO" if ativo else "DEVOLVIDO",
            "substituicoes": [
                {"data": t["data"], "placa": t["placa"]}
                for t in substituicoes_da_linha(linha)
                if t["placa"] in placas
            ],
        })
    return resultado


def extrair_tabelas_de_preco() -> list[dict]:
    """
    Lê as grades da Unidas e da Localiza, que convivem lado a lado na mesma aba.

    A Unidas ocupa as colunas A–F e a Localiza as colunas H–L, cada uma com seu
    próprio conjunto de pacotes de KM — a razão pela qual o modelo guarda pacotes
    como linhas, e não como colunas.
    """
    wb = abrir("Valores Locação Mensal_Unidas e Localiza.xlsx")
    tabelas = []

    for aba, ano in (("Valores Locadoras 2025", 2025), ("Valores Locadoras 2026", 2026)):
        if aba not in wb.sheetnames:
            continue
        linhas = list(wb[aba].iter_rows(min_row=2, max_col=12, values_only=True))
        cabecalho = linhas[0]
        # Os pacotes vêm no cabeçalho ora como número, ora como texto ("3000"),
        # conforme a célula foi digitada — daí a conversão tolerante.
        pacotes_unidas = pacotes_do_cabecalho(cabecalho[2:6])
        pacotes_localiza = pacotes_do_cabecalho(cabecalho[9:12])

        for locadora, coluna_grupo, pacotes, deslocamento in (
            ("Unidas", 0, pacotes_unidas, 2),
            ("Localiza", 7, pacotes_localiza, 9),
        ):
            grupos = []
            vistos: set[str] = set()
            for linha in linhas[1:]:
                codigo = texto(linha[coluna_grupo])
                veiculos = texto(linha[coluna_grupo + 1])
                if not codigo or not veiculos:
                    continue
                itens = []
                for i, pacote in enumerate(pacotes):
                    valor = moeda(linha[deslocamento + i]) if deslocamento + i < len(linha) else None
                    # A aba de 2026 repete a grade mais abaixo com o **percentual de
                    # reajuste** (0,0267 = 2,67%) em vez do preço. Um aluguel mensal
                    # nunca custa menos de R$ 100, o que separa os dois blocos sem
                    # depender da posição das linhas.
                    if valor and valor >= 100:
                        itens.append({"pacoteKm": pacote, "valorMensal": valor})
                if not itens or codigo.upper() in vistos:
                    continue
                vistos.add(codigo.upper())
                if itens:
                    grupos.append({
                        "codigo": codigo,
                        "veiculosDoGrupo": veiculos[:300],
                        "categoria": categoria_do_grupo(veiculos),
                        "pacotes": itens,
                    })
            if grupos:
                tabelas.append({
                    "locadoraNome": locadora,
                    "anoVigencia": ano,
                    "grupos": grupos,
                    "kmExcedente": km_excedente_de(locadora),
                })
    wb.close()
    return tabelas


def pacotes_do_cabecalho(celulas) -> list[int]:
    pacotes = []
    for celula in celulas:
        digitos = re.sub(r"\D", "", texto(celula))
        if digitos:
            pacotes.append(int(digitos))
    return pacotes


def categoria_do_grupo(veiculos: str) -> str:
    alvo = sem_acento(veiculos).upper()
    if any(p in alvo for p in ("S10", "L200", "HILUX", "RANGER", "AMAROK")):
        return "QUATRO_X_QUATRO"
    if any(p in alvo for p in ("RENEGADE", "T-CROSS", "TCROSS", "KICKS", "CACTUS", "CRETA", "TORO", "COMPASS")):
        return "SUV"
    if any(p in alvo for p in ("FIORINO", "STRADA", "SAVEIRO", "DOBLO", "SPIN", "PARTNER")):
        return "UTILITARIO"
    return "PASSEIO"


def km_excedente_de(locadora: str) -> list[dict]:
    """
    Valores da aba "Cotação Nacional".

    A Unidas cobra um valor por categoria, igual para todos os pacotes; a Localiza
    diferencia por pacote. O modelo acomoda os dois com `pacoteKm` opcional (RN-06).
    """
    if locadora == "Unidas":
        return [
            {"categoria": "PASSEIO", "pacoteKm": None, "valorKm": 0.60},
            {"categoria": "UTILITARIO", "pacoteKm": None, "valorKm": 0.60},
            {"categoria": "SUV", "pacoteKm": None, "valorKm": 1.20},
            {"categoria": "QUATRO_X_QUATRO", "pacoteKm": None, "valorKm": 3.20},
        ]
    return [
        {"categoria": "PASSEIO", "pacoteKm": 3000, "valorKm": 0.50},
        {"categoria": "PASSEIO", "pacoteKm": 4000, "valorKm": 1.00},
        {"categoria": "PASSEIO", "pacoteKm": 5000, "valorKm": 1.00},
        {"categoria": "SUV", "pacoteKm": 3000, "valorKm": 1.00},
        {"categoria": "SUV", "pacoteKm": 4000, "valorKm": 2.00},
        {"categoria": "SUV", "pacoteKm": 5000, "valorKm": 2.00},
        {"categoria": "QUATRO_X_QUATRO", "pacoteKm": 3000, "valorKm": 3.00},
        {"categoria": "QUATRO_X_QUATRO", "pacoteKm": 4000, "valorKm": 6.00},
        {"categoria": "QUATRO_X_QUATRO", "pacoteKm": 5000, "valorKm": 6.00},
    ]


def atribuir_grupos_tarifarios(veiculos: list[dict], tabelas: list[dict]) -> None:
    """
    Liga cada veículo ao grupo tarifário da sua locadora.

    Sem esse vínculo, o fechamento mensal não sabe qual valor contratado aplicar
    (RN-06) — e o painel não consegue estimar o custo da frota.
    """
    grades: dict[str, list[dict]] = {}
    for tabela in tabelas:
        if tabela["anoVigencia"] == max(t["anoVigencia"] for t in tabelas):
            grades[tabela["locadoraNome"]] = tabela["grupos"]

    for veiculo in veiculos:
        grupos = grades.get(veiculo["locadoraNome"])
        if not grupos:
            continue
        modelo = sem_acento(veiculo["modelo"]).upper()
        # Primeiro tenta casar o modelo com a descrição do grupo; se não houver
        # correspondência, cai no primeiro grupo da mesma categoria.
        escolhido = next(
            (g for g in grupos if any(
                termo and termo in modelo
                for termo in (sem_acento(p).upper().strip() for p in re.split(r"[,/]", g["veiculosDoGrupo"]))
            )),
            None,
        )
        if not escolhido:
            escolhido = next((g for g in grupos if g["categoria"] == veiculo["categoria"]), None)
        if escolhido:
            veiculo["grupoTarifario"] = escolhido["codigo"]


def main() -> None:
    contratos = extrair_contratos()
    obras = extrair_obras(contratos)
    locadoras = extrair_locadoras()
    condutores = extrair_condutores(contratos, obras)
    veiculos, placas_rejeitadas = extrair_veiculos(contratos, locadoras)
    fornecedores = extrair_fornecedores({o["codigo"] for o in obras})
    tabelas = extrair_tabelas_de_preco()
    atribuir_grupos_tarifarios(veiculos, tabelas)
    contratos_de_locacao = extrair_contratos_de_locacao(
        contratos,
        {o["codigo"] for o in obras},
        {v["placa"] for v in veiculos},
        condutores,
    )

    acervo = {
        "origem": "Planilhas legadas em arquivos/ — extraídas por scripts/extrair-acervo.py",
        "aviso": (
            "Dados de desenvolvimento. Os CPFs dos condutores e as validades de CNH são "
            "sintéticos, porque as planilhas não os contêm; as credenciais de portal são "
            "de demonstração. Nada disso deve existir em produção."
        ),
        "obras": obras,
        "locadoras": locadoras,
        "condutores": condutores,
        "veiculos": veiculos,
        "fornecedores": fornecedores,
        "tabelasDePreco": tabelas,
        "contratosDeLocacao": contratos_de_locacao,
    }

    SAIDA.parent.mkdir(parents=True, exist_ok=True)
    SAIDA.write_text(json.dumps(acervo, ensure_ascii=False, indent=1), encoding="utf-8")

    print(f"Acervo gravado em {SAIDA.relative_to(RAIZ)}")
    print(f"  obras .............. {len(obras)}")
    print(f"  locadoras .......... {len(locadoras)}")
    print(f"  condutores ......... {len(condutores)}")
    print(f"  veículos ........... {len(veiculos)}")
    print(f"  fornecedores ....... {len(fornecedores)}")
    print(f"  tabelas de preço ... {len(tabelas)}")
    ativos = sum(1 for c in contratos_de_locacao if c["status"] == "ATIVO")
    trocas = sum(len(c["substituicoes"]) for c in contratos_de_locacao)
    com_troca = sum(1 for c in contratos_de_locacao if c["substituicoes"])
    print(f"  contratos .......... {len(contratos_de_locacao)} ({ativos} ativos)")
    print(f"  substituições ...... {trocas} em {com_troca} contratos")
    if placas_rejeitadas:
        print(f"  placas descartadas . {len(placas_rejeitadas)} → {placas_rejeitadas[:5]}")


if __name__ == "__main__":
    main()
