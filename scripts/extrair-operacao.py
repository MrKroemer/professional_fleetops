#!/usr/bin/env python3
"""
Extrai os lançamentos da operação mensal das planilhas por obra.

Os arquivos de `Controles por Obra` têm uma aba por condutor, e dentro dela um bloco de
quatro linhas por mês. O bloco reúne três controles lado a lado, em faixas de colunas:

    A–H   REGISTRO DE KM   quatro leituras de hodômetro no mês (1ª a 4ª)
    J–V   ABASTECIMENTO    pares DATA/VALOR, um por nota
    W–AD  LAVA_RAPIDO      pares DATA/VALOR
    AE+   BORRACHARIA      pares DATA/VALOR

Sobre a quilometragem: a planilha guarda **leituras**, não trechos. Não há a data de cada
leitura — só o mês. Por isso cada mês vira **um** registro de KM, com o hodômetro da
primeira leitura preenchida como inicial e o da última como final, datado no fim do mês.

É a leitura honesta do que existe: inventar quatro datas semanais produziria um histórico
com precisão que o dado não tem, e a distância mensal — que é o que o fechamento e a
RN-06 usam — sai idêntica de qualquer forma.

Uso:  python3 scripts/extrair-operacao.py
Saída: backend/src/main/resources/db/seed/operacao.json
"""

import json
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parent.parent
PASTA = RAIZ / "arquivos" / "Controles por Obra - KM, Lava-jato, Borracharia"
SAIDA = RAIZ / "backend/src/main/resources/db/seed/operacao.json"

MESES = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "abril": 4, "maio": 5, "junho": 6,
    "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
}

# Faixas de coluna de cada controle, em índice base zero.
COLUNAS_LEITURAS = range(1, 5)      # B..E
FAIXA_ABASTECIMENTO = range(10, 22)  # K..V
FAIXA_LAVA_JATO = range(23, 31)      # X..AD
FAIXA_BORRACHARIA = range(31, 40)    # AF..

PADRAO_PLACA = re.compile(r"^[A-Z]{3}[0-9][A-Z0-9][0-9]{2}$")

# Uma nota de combustível fora desta faixa é quase certamente outra coisa na célula —
# um total anual, um código. O teto vem do maior tanque da frota a preço de mercado.
VALOR_MINIMO, VALOR_MAXIMO = 20.0, 2000.0


def sem_acento(texto: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn"
    ).lower().strip()


def normalizar_placa(valor) -> str | None:
    if not valor:
        return None
    limpa = re.sub(r"[^A-Za-z0-9]", "", str(valor)).upper()
    return limpa if PADRAO_PLACA.match(limpa) else None


def como_data(valor) -> date | None:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return None


def como_valor(valor) -> float | None:
    """Aceita só números plausíveis para uma nota — o resto é ruído de célula."""
    if not isinstance(valor, (int, float)) or isinstance(valor, bool):
        return None
    if VALOR_MINIMO <= float(valor) <= VALOR_MAXIMO:
        return round(float(valor), 2)
    return None


def pares_de_lancamento(cabecalho, valores, faixa) -> list[dict]:
    """
    Lê os pares DATA/VALOR de uma faixa de colunas.

    A data está na linha de cabeçalho do bloco e o valor na linha seguinte, na mesma
    coluna. Um par só entra quando as duas metades existem: uma data sem valor é uma
    célula que alguém começou a preencher, e um valor sem data não tem onde ser lançado.
    """
    lancamentos = []
    for coluna in faixa:
        if coluna >= len(cabecalho) or coluna >= len(valores):
            continue
        quando = como_data(cabecalho[coluna])
        quanto = como_valor(valores[coluna])
        if quando and quanto:
            lancamentos.append({"data": quando.isoformat(), "valor": quanto})
    return lancamentos


def separar_series(registros: list[dict]) -> list[list[dict]]:
    """
    Quebra os registros onde o hodômetro despenca.

    Uma queda de mais de 5.000 km entre um mês e o seguinte não é erro de digitação: é
    outro carro. A planilha registra a substituição no cabeçalho da aba, e sem esta
    separação todos os meses posteriores seriam recusados pela RN-03 — perdendo justamente
    a operação do veículo substituto.
    """
    QUEDA_DE_TROCA = 5000
    ordenados = sorted(registros, key=lambda r: r["data"])
    series: list[list[dict]] = []
    atual: list[dict] = []
    for registro in ordenados:
        if atual and registro["kmInicial"] < atual[-1]["kmFinal"] - QUEDA_DE_TROCA:
            series.append(atual)
            atual = []
        atual.append(registro)
    if atual:
        series.append(atual)
    return series


def extrair_aba(ws, nome_da_aba: str) -> list[dict] | None:
    """Extrai o cabeçalho e todos os blocos mensais preenchidos de uma aba de condutor."""
    linhas = list(ws.iter_rows(values_only=True))
    if len(linhas) < 12:
        return None

    def celula(linha, coluna):
        return linhas[linha][coluna] if linha < len(linhas) and coluna < len(linhas[linha]) else None

    placa = normalizar_placa(celula(1, 4)) or normalizar_placa(celula(1, 3))
    if not placa:
        return None

    # A linha 3 registra a substituição: data, placa nova e modelo. Quando o hodômetro
    # despenca no meio da aba, é este o veículo que passou a rodar.
    placa_substituta = normalizar_placa(celula(2, 4)) or normalizar_placa(celula(2, 3))

    # O ano aparece na coluna A apenas no bloco em que ele VIRA — tipicamente em Janeiro.
    # Os blocos acima do primeiro marcador pertencem ao ano anterior.
    #
    # Aplicar o marcador a todos os blocos, como esta função fazia antes, jogava Novembro
    # e Dezembro para o ano seguinte. O hodômetro então "voltava" ao ordenar por data, e
    # a RN-03 recusava 135 transições que na planilha estavam perfeitamente em ordem.
    marcadores = []
    for i, linha in enumerate(linhas):
        marcador = linha[0] if linha else None
        if isinstance(marcador, (int, float)) and not isinstance(marcador, bool) and 2020 <= marcador <= 2035:
            marcadores.append((i, int(marcador)))
    ano_inicial = (marcadores[0][1] - 1) if marcadores else 2025

    registros, abastecimentos, servicos = [], [], []

    for i, linha in enumerate(linhas):
        rotulo = celula(i, 1)
        if not isinstance(rotulo, str):
            continue
        mes = MESES.get(sem_acento(rotulo))
        if not mes:
            continue

        cabecalho = linhas[i + 1] if i + 1 < len(linhas) else ()
        valores = linhas[i + 2] if i + 2 < len(linhas) else ()

        # O ano vem, de preferência, das datas de abastecimento do próprio bloco: elas são
        # datas de verdade, com ano e mês, e concordam com o rótulo do mês por construção.
        # É a âncora mais confiável — o marcador da coluna A falta em várias abas, e onde
        # falta todos os meses cairiam no mesmo ano, produzindo dois "Novembro" iguais.
        ano = None
        for coluna in list(FAIXA_ABASTECIMENTO) + list(FAIXA_LAVA_JATO):
            quando = como_data(cabecalho[coluna]) if coluna < len(cabecalho) else None
            if quando and quando.month == mes:
                ano = quando.year
                break

        if ano is None:
            # Sem datas no bloco, resta o marcador: vale do ponto em que aparece para
            # baixo, e o que está acima dele pertence ao ano anterior.
            ano = ano_inicial
            for posicao, valor in marcadores:
                if posicao <= i:
                    ano = valor

        leituras = [
            int(valores[c]) for c in COLUNAS_LEITURAS
            if c < len(valores) and isinstance(valores[c], (int, float))
            and not isinstance(valores[c], bool) and valores[c] > 0
        ]
        if len(leituras) >= 2 and leituras[-1] >= leituras[0]:
            ultimo_dia = (date(ano + (mes == 12), (mes % 12) + 1, 1) - __import__("datetime").timedelta(days=1))
            registros.append({
                "data": ultimo_dia.isoformat(),
                "kmInicial": leituras[0],
                "kmFinal": leituras[-1],
                "observacao": "Leituras de hodômetro do mês, consolidadas da planilha da obra.",
            })

        abastecimentos += pares_de_lancamento(cabecalho, valores, FAIXA_ABASTECIMENTO)
        for servico in pares_de_lancamento(cabecalho, valores, FAIXA_LAVA_JATO):
            servicos.append({**servico, "tipo": "LAVA_JATO"})
        for servico in pares_de_lancamento(cabecalho, valores, FAIXA_BORRACHARIA):
            servicos.append({**servico, "tipo": "BORRACHARIA"})

    if not (registros or abastecimentos or servicos):
        return None

    condutor = nome_da_aba.split("-")[0].strip()
    series = separar_series(registros)

    # A primeira série é do veículo do cabeçalho; a seguinte, do substituto. Os
    # abastecimentos e serviços ficam com o veículo original: a planilha não os separa
    # por veículo, e dividi-los por data seria inventar um critério que ela não tem.
    blocos = [{
        "placa": placa,
        "condutor": condutor,
        "registrosDeKm": series[0] if series else [],
        "abastecimentos": abastecimentos,
        "servicos": servicos,
    }]
    if len(series) > 1 and placa_substituta:
        blocos.append({
            "placa": placa_substituta,
            "condutor": condutor,
            "registrosDeKm": [r for serie in series[1:] for r in serie],
            "abastecimentos": [],
            "servicos": [],
        })
    return blocos


def main() -> None:
    blocos = []
    for arquivo in sorted(PASTA.glob("*.xlsx")):
        wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
        for aba in wb.sheetnames:
            if aba in ("RESUMO", "Controle Central") or aba.startswith("CAR_"):
                continue
            extraidos = extrair_aba(wb[aba], aba)
            for extraido in extraidos or []:
                extraido["obraArquivo"] = arquivo.stem
                blocos.append(extraido)
        wb.close()

    total_km = sum(len(b["registrosDeKm"]) for b in blocos)
    total_abast = sum(len(b["abastecimentos"]) for b in blocos)
    total_serv = sum(len(b["servicos"]) for b in blocos)

    SAIDA.parent.mkdir(parents=True, exist_ok=True)
    SAIDA.write_text(json.dumps(blocos, ensure_ascii=False, indent=1), encoding="utf-8")

    print(f"{len(blocos)} veículos com operação")
    print(f"  {total_km} registros de KM")
    print(f"  {total_abast} abastecimentos")
    print(f"  {total_serv} serviços")
    print(f"→ {SAIDA.relative_to(RAIZ)}")


if __name__ == "__main__":
    main()
